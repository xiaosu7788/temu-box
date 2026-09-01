from openpyxl import Workbook, load_workbook

from app.services.orders import build_delivery_sku_map, generate_summary


def test_delivery_to_summary_pipeline(tmp_path):
    delivery = Workbook()
    delivery_ws = delivery.active
    delivery_ws.append(["", "参考单号"] + [""] * 17 + ["SKU", "数量"])
    delivery_ws.append(["", "PO-TEST-001"] + [""] * 17 + ["MB131-491", 2])
    delivery_path = tmp_path / "delivery.xlsx"
    delivery.save(delivery_path)

    sales = Workbook()
    sales_ws = sales.active
    sales_ws.append(["", "", "PO单号", "", "", "结算金额"])
    sales_ws.append(["", "", "PO-TEST-001", "", "", 88.5])
    sales_path = tmp_path / "sales.xlsx"
    sales.save(sales_path)

    output_path = tmp_path / "summary.xlsx"
    po_map = build_delivery_sku_map(delivery_path)
    stats = generate_summary(
        sales_path,
        {"MB131-491": {"price": 17.1, "set_type": "6件套"}},
        po_map,
        output_path,
        {"MB131-491": "6件套"},
    )

    assert stats["matched"] == 1
    assert stats["unmatched"] == 0
    workbook = load_workbook(output_path, data_only=True)
    row = [workbook.active.cell(2, column).value for column in range(1, 8)]
    workbook.close()
    assert row == ["PO-TEST-001", "MB131-491", 17.1, 2, 48.2, 88.5, "6件套"]
