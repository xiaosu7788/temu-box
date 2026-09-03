import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.database import create_user
from app.main import app
from app.services.activity import activity_base_price, parse_skc, preview_activity_workbook, process_activity_workbook
from app.services.auth import hash_password


def make_activity_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "活动申报价格"
    sheet.append(["活动类型(活动主题）", "SPU ID", "SKC ID", "SKC货号", "SKU ID", "活动申报价格"])
    sheet.append(["活动", "1", "1", "MB131-A-5", "10", 17])
    sheet.append(["活动", "2", "2", "MB131-B-5", "11", 18])
    sheet.append(["活动", "3", "3", "MB131-C-5", "12", 16])
    sheet.append(["活动", "4", "4", "y1-4piece", "13", 42])
    sheet.append(["活动", "5", "5", "y1-5piece", "14", 45.5])
    sheet.append(["活动", "6", "6", "y1-6piece", "15", 47])

    inventory = workbook.create_sheet("活动库存")
    inventory.append(["活动类型(活动主题）", "SPU ID", "活动库存"])
    inventory.append(["活动", "1", 15])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_parse_activity_skc_rules():
    assert parse_skc("MB131-A-5") == ("single", 5.0)
    assert parse_skc("y1-8piece") == ("set", 8.0)
    assert activity_base_price(("single", 5.0)) == 17.0
    assert activity_base_price(("set", 12.0)) == 92.0


def test_parse_custom_activity_skc_rules():
    common = {
        "set_keywords": ["piece", "件套", "套装"],
        "set_mappings": [{"pattern": "四件组合", "pieces": 4}],
        "single_mode": "first_segment",
        "single_delimiter": "-",
        "single_marker": "price",
    }
    assert parse_skc("y1-4piece", common) == ("set", 4.0)
    assert parse_skc("ABC-8件套", common) == ("set", 8.0)
    assert parse_skc("春季四件组合-A", common) == ("set", 4.0)
    assert parse_skc("5-MB131-A", common) == ("single", 5.0)

    last_segment = {**common, "single_mode": "last_segment"}
    assert parse_skc("MB131-A-5.5", last_segment) == ("single", 5.5)

    after_marker = {**common, "single_mode": "after_marker"}
    assert parse_skc("MB131-price17.1", after_marker) == ("single", 17.1)

    empty_set_marker = {**common, "set_keywords": [""]}
    assert parse_skc("SET-A-10", empty_set_marker) == ("set", 10.0)


def test_activity_price_uses_admin_settings():
    settings = {"activity": {"headcost": 6, "operation_fee": 8, "set_prices": {"4": 50}, "single_tiers": [{"min_price": 0, "profit": 0}, {"min_price": 15, "profit": 4}]}}
    assert activity_base_price(("single", 15.0), settings) == 33.0
    assert activity_base_price(("set", 4.0), settings) == 50.0


def test_process_activity_workbook_updates_filters_and_preserves_sheets(tmp_path):
    output = tmp_path / "result.xlsx"
    stats = process_activity_workbook(make_activity_workbook(), output)

    assert stats["input_data_rows"] == 6
    assert stats["processed_rows"] == 6
    assert stats["unchanged_rows"] == 2
    assert stats["removed_rows"] == 2
    assert stats["updated_rows"] == 2
    assert stats["remaining_data_rows"] == 4

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["活动申报价格", "活动库存"]
    sheet = workbook["活动申报价格"]
    rows = [(sheet.cell(row, 4).value, sheet.cell(row, 6).value) for row in range(2, sheet.max_row + 1)]
    workbook.close()

    assert rows[0] == ("MB131-A-5", 17)
    assert rows[1][0] == "MB131-B-5"
    assert 17 < rows[1][1] <= 18
    assert rows[2] == ("y1-4piece", 42)
    assert rows[3][0] == "y1-5piece"
    assert 45 < rows[3][1] <= 45.5


def test_activity_uses_default_skc_rules_from_settings(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["MB131-price17.1", 40])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    settings = {
        "activity": {
            "default_skc_rules": {
                "set_keywords": ["piece"],
                "set_mappings": [],
                "single_mode": "after_marker",
                "single_delimiter": "-",
                "single_marker": "price",
            }
        }
    }
    output = tmp_path / "default-rules.xlsx"

    stats = process_activity_workbook(content.getvalue(), output, settings)

    assert stats["processed_rows"] == 1
    assert stats["custom_skc_rules"] is False

def test_activity_uplift_limit_uses_configured_value(tmp_path):
    output = tmp_path / "limited-result.xlsx"
    settings = {"activity": {"uplift_limit": 0.25}}
    stats = process_activity_workbook(make_activity_workbook(), output, settings)

    workbook = load_workbook(output, data_only=True)
    sheet = workbook["活动申报价格"]
    rows = [(sheet.cell(row, 4).value, sheet.cell(row, 6).value) for row in range(2, sheet.max_row + 1)]
    workbook.close()

    assert stats["uplift_limit"] == 0.25
    assert rows[1][0] == "MB131-B-5"
    assert 17 < rows[1][1] <= 17.25


def test_preview_custom_activity_rules():
    rules = {
        "set_keywords": ["piece", "件套"],
        "set_mappings": [],
        "single_mode": "first_segment",
        "single_delimiter": "-",
        "single_marker": "price",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["5-MB131-A", 20])
    sheet.append(["ABC-8件套", 80])
    sheet.append(["UNKNOWN", 20])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    preview = preview_activity_workbook(content.getvalue(), rules)
    assert preview["total_rows"] == 3
    assert preview["single_rows"] == 1
    assert preview["set_rows"] == 1
    assert preview["unrecognized_rows"] == 1
    assert preview["items"][0]["base_price"] == 17


def test_activity_preview_endpoint_uses_custom_rules():
    create_user("activity_preview_user", hash_password("previewpass123"), status="approved")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["MB131-price17.1", 30])
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    rules = {
        "set_keywords": ["piece"],
        "set_mappings": [],
        "single_mode": "after_marker",
        "single_delimiter": "-",
        "single_marker": "price",
    }

    with TestClient(app) as client:
        client.post("/api/auth/login", json={"username": "activity_preview_user", "password": "previewpass123"})
        response = client.post(
            "/api/activities/preview",
            data={"skc_rules": json.dumps(rules, ensure_ascii=False)},
            files={"file": ("preview.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["items"][0]["result"] == "单品"
        assert response.json()["items"][0]["value"] == 17.1


def test_activity_task_is_visible_after_submission():
    create_user("activity_owner", hash_password("activitypass123"), status="approved")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["MB131-A-5", 17])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with TestClient(app) as client:
        login = client.post("/api/auth/login", json={"username": "activity_owner", "password": "activitypass123"})
        assert login.status_code == 200
        response = client.post(
            "/api/activities/bulk",
            files={"file": ("activity.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 202
        job_id = response.json()["id"]
        listed = client.get("/api/activities")
        assert listed.status_code == 200
        assert job_id in {item["id"] for item in listed.json()["items"]}


def test_custom_uplift_does_not_change_default_settings():
    create_user("custom_uplift_user", hash_password("custompass123"), status="approved")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["MB131-CUSTOM-5", 18])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with TestClient(app) as client:
        client.post("/api/auth/login", json={"username": "custom_uplift_user", "password": "custompass123"})
        default_before = client.get("/api/settings").json()["activity"]["uplift_limit"]
        response = client.post(
            "/api/activities/bulk",
            data={"uplift_limit": "0.25"},
            files={"file": ("custom-uplift.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 202
        assert client.get("/api/settings").json()["activity"]["uplift_limit"] == default_before


def test_admin_frontend_activity_list_is_scoped_but_admin_list_includes_all_jobs():
    create_user("activity_admin", hash_password("activityadmin123"), role="admin", status="approved")
    create_user("activity_user", hash_password("activityuser123"), status="approved")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SKC货号", "活动申报价格"])
    sheet.append(["MB131-ADMIN-5", 17])
    content = BytesIO()
    workbook.save(content)
    workbook.close()

    with TestClient(app) as client:
        client.post("/api/auth/login", json={"username": "activity_user", "password": "activityuser123"})
        response = client.post(
            "/api/activities/bulk",
            files={"file": ("activity-user.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 202
        job_id = response.json()["id"]

        client.post("/api/auth/logout")
        client.post("/api/auth/login", json={"username": "activity_admin", "password": "activityadmin123"})
        frontend_ids = {item["id"] for item in client.get("/api/activities").json()["items"]}
        admin_ids = {item["id"] for item in client.get("/api/admin/activity-tasks").json()["items"]}
        assert job_id not in frontend_ids
        assert job_id in admin_ids
        assert client.get(f"/api/activities/{job_id}").status_code == 404
        assert client.delete(f"/api/activities/{job_id}").status_code == 404
