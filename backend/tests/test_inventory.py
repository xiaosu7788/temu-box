from openpyxl import Workbook
from fastapi.testclient import TestClient

from app.main import app
from app.services.auth import admin_user
from app.database import create_inventory_item, update_inventory_item, get_inventory_catalog, save_inventory_catalog
from app.services.inventory import build_price_catalog


def test_price_header_can_appear_in_middle(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存0"
    sheet.cell(30, 12, "价格")
    sheet.cell(31, 8, "MB131-TEST1")
    sheet.cell(31, 9, 99)
    sheet.cell(31, 12, 17.1)
    path = tmp_path / "header.xlsx"
    workbook.save(path)

    catalog = build_price_catalog(path)

    assert catalog["MB131-TEST1"]["price"] == 17.1
    assert catalog["MB131-TEST1"]["source_column"] == 12


def test_price_search_expands_around_sku(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存0"
    sheet.cell(1, 8, "MB131-TEST2")
    sheet.cell(1, 9, "六件套")
    sheet.cell(1, 10, 18.2)
    path = tmp_path / "sides.xlsx"
    workbook.save(path)

    catalog = build_price_catalog(path)

    assert catalog["MB131-TEST2"]["price"] == 18.2
    assert catalog["MB131-TEST2"]["set_type"] == "6件套"


def test_duplicate_sku_uses_lowest_price(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存0"
    sheet.cell(1, 8, "MB131-TEST3")
    sheet.cell(1, 9, 20)
    sheet.cell(2, 8, "MB131-TEST3")
    sheet.cell(2, 9, 15)
    path = tmp_path / "duplicates.xlsx"
    workbook.save(path)

    catalog = build_price_catalog(path)

    assert catalog["MB131-TEST3"]["price"] == 15


def test_inventory_items_endpoint_filters_and_paginates(monkeypatch):
    from app.services.auth import current_user

    app.dependency_overrides[current_user] = lambda: {"id": 1, "role": "user", "status": "approved"}
    monkeypatch.setattr(
        "app.main.load_price_catalog",
        lambda: {
            "MB131-A": {"sku": "MB131-A", "price": 17.1, "set_type": "单品"},
            "MB131-B": {"sku": "MB131-B", "price": 22.0, "set_type": "2件套"},
        },
    )

    try:
        response = TestClient(app).get(
            "/api/inventory/items",
            params={"query": "mb131-b", "page": 1, "page_size": 10},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.json() == {
        "total": 1,
        "items": [{"sku": "MB131-B", "price": 22.0, "set_type": "2件套"}],
    }


def test_inventory_item_can_be_created_and_updated():
    signature = {"path": "manual-inventory.xlsx", "size": 1, "mtime_ns": 1, "parser_version": 3}
    save_inventory_catalog(signature, {})

    created = create_inventory_item("MB131-MANUAL", 17.1, "单品")
    assert created["sku"] == "MB131-MANUAL"
    assert created["source_sheet"] == "手动维护"
    assert get_inventory_catalog()["MB131-MANUAL"]["price"] == 17.1

    updated = update_inventory_item("MB131-MANUAL", "MB131-MANUAL-2", 18.2, "6件套")
    assert updated["sku"] == "MB131-MANUAL-2"
    assert updated["set_type"] == "6件套"
    assert "MB131-MANUAL" not in get_inventory_catalog()
    assert get_inventory_catalog()["MB131-MANUAL-2"]["price"] == 18.2

def test_admin_can_create_and_update_inventory_item_but_user_cannot():
    from app.database import create_user
    from app.services.auth import hash_password

    admin = create_user("inventory_admin", hash_password("inventoryadmin123"), role="admin", status="approved")
    user = create_user("inventory_user", hash_password("inventoryuser123"), status="approved")
    with TestClient(app) as client:
        assert client.post("/api/auth/login", json={"username": "inventory_user", "password": "inventoryuser123"}).status_code == 200
        assert client.post("/api/admin/inventory/items", json={"sku": "MB131-DENIED", "price": 1, "set_type": "单品"}).status_code == 403
        client.post("/api/auth/logout")
        assert client.post("/api/auth/login", json={"username": "inventory_admin", "password": "inventoryadmin123"}).status_code == 200
        created = client.post("/api/admin/inventory/items", json={"sku": "mb131-api", "price": 19.5, "set_type": "单品"})
        assert created.status_code == 201
        assert created.json()["item"]["sku"] == "MB131-API"
        updated = client.put("/api/admin/inventory/items/MB131-API", json={"sku": "MB131-API", "price": 20, "set_type": "4件套"})
        assert updated.status_code == 200
        assert updated.json()["item"]["set_type"] == "4件套"
