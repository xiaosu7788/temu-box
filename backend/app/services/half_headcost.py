from __future__ import annotations

import io
import json
import os
import re
import threading
import time
from pathlib import Path
from typing import Dict, Union

from openpyxl import load_workbook

from app.config import HALF_HEADCOST_PATH, HALF_HEADCOST_SEED_PATH
from app.database import (
    delete_half_entry,
    load_half_entries,
    merge_half_entries,
)
from app.services.inventory import CODE_RE, extract_set_type


SKU_HEADER_RE = re.compile(
    r"(sku|货号|商品编号|商品编码|产品编号|产品编码)", re.IGNORECASE
)
_LOCK = threading.Lock()


def extract_sku_types(source: Union[Path, bytes, bytearray]) -> Dict[str, str]:
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True, keep_links=False)
    else:
        wb = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    sku_types = {}
    try:
        for ws in wb.worksheets:
            sku_columns = set()
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                for index, value in enumerate(row):
                    if value is not None and SKU_HEADER_RE.search(str(value)):
                        sku_columns.add(index)
            for row in ws.iter_rows(values_only=True):
                texts = [
                    str(value) for value in row
                    if value is not None and not str(value).startswith("=DISPIMG")
                ]
                set_type = extract_set_type(texts) or "单品"
                values = [row[index] for index in sku_columns if index < len(row)] if sku_columns else row
                for value in values:
                    if value is None or str(value).startswith("="):
                        continue
                    for sku in CODE_RE.findall(str(value).strip()):
                        sku_types[sku] = set_type
    finally:
        wb.close()
    return sku_types


def _read() -> Dict[str, str]:
    try:
        with HALF_HEADCOST_PATH.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        values = data.get("sku_types", data) if isinstance(data, dict) else {}
        return {str(key): str(value) for key, value in values.items()}
    except (FileNotFoundError, json.JSONDecodeError, OSError, TypeError):
        return {}


def _write(values: Dict[str, str]) -> None:
    HALF_HEADCOST_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = HALF_HEADCOST_PATH.with_suffix(HALF_HEADCOST_PATH.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "sku_types": dict(sorted(values.items())),
                "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            },
            handle,
            ensure_ascii=False,
            indent=2,
        )
    os.replace(temp_path, HALF_HEADCOST_PATH)


def load_entries() -> Dict[str, str]:
    with _LOCK:
        values = load_half_entries()
        if not values and HALF_HEADCOST_SEED_PATH.exists():
            seeds = extract_sku_types(HALF_HEADCOST_SEED_PATH)
            merge_half_entries(seeds)
            values = load_half_entries()
            _write(values)
        return values


def merge_upload(source: Union[Path, bytes, bytearray]) -> dict:
    incoming = extract_sku_types(source)
    if not incoming:
        raise ValueError("上传表格中未识别到 MB131- 开头的 SKU")
    with _LOCK:
        added, total = merge_half_entries(incoming)
        values = load_half_entries()
        _write(values)
    return {
        "incoming": len(incoming),
        "added": added,
        "total": total,
    }


def delete_entry(sku: str) -> bool:
    with _LOCK:
        existed = delete_half_entry(sku)
        if existed:
            _write(load_half_entries())
        return existed
