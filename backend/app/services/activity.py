from __future__ import annotations

import io
import random
import re
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook


ACTIVITY_PRICE_BASE = {
    4: 42.0,
    5: 45.0,
    6: 48.0,
    8: 71.0,
    10: 75.0,
    12: 92.0,
}
SINGLE_SKC_RE = re.compile(r"-([0-9]+(?:\.[0-9]+)?)\s*$")
SET_SKC_RE = re.compile(r"^y\d+\s*-\s*(\d+)\s*piece\s*$", re.IGNORECASE)
HEADER_NAMES = {
    "skc": "SKC货号",
    "price": "活动申报价格",
    "spu_id": "SPU ID",
    "skc_id": "SKC ID",
    "sku_id": "SKU ID",
}
REQUIRED_HEADER_KEYS = {"skc", "price"}
ID_RULE_TYPES = ("SPU", "SKC", "SKU")
SINGLE_PARSE_MODES = {"first_segment", "last_segment", "after_marker"}
MAX_PREVIEW_ROWS = 100
MAX_ID_PROFIT_RULES = 1000


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _find_headers(workbook) -> Tuple[object, int, Dict[str, int]]:
    for worksheet in workbook.worksheets:
        for row in range(1, min(worksheet.max_row, 30) + 1):
            columns = {}
            for column in range(1, worksheet.max_column + 1):
                value = _text(worksheet.cell(row, column).value)
                for key, header in HEADER_NAMES.items():
                    if value == header:
                        columns[key] = column
            if REQUIRED_HEADER_KEYS.issubset(columns):
                return worksheet, row, columns
    raise ValueError("未找到同时包含“SKC货号”和“活动申报价格”的工作表")


def _identifier_key(value: object) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value).casefold()


def normalize_id_profit_rules(value: object) -> list[dict]:
    if value is None:
        return []
    if not isinstance(value, list) or len(value) > MAX_ID_PROFIT_RULES:
        raise ValueError(f"ID利润条件最多可设置{MAX_ID_PROFIT_RULES}条")

    normalized = []
    seen = set()
    for item in value:
        if not isinstance(item, dict):
            raise ValueError("ID利润条件格式不正确")
        id_type = _text(item.get("id_type")).upper()
        if id_type not in ID_RULE_TYPES:
            raise ValueError("ID类型只能是SPU、SKC或SKU")
        identifier = _text(item.get("id"))
        if not identifier or len(identifier) > 120:
            raise ValueError("ID不能为空且不能超过120个字符")
        key = (id_type, _identifier_key(identifier))
        if key in seen:
            raise ValueError(f"{id_type} ID“{identifier}”不能重复设置")
        seen.add(key)
        try:
            profit = float(item.get("profit", 0))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{id_type} ID“{identifier}”的利润调整值不正确") from exc
        if profit < -100000 or profit > 100000:
            raise ValueError("ID利润调整值必须在-100000到100000之间")
        normalized.append({"id_type": id_type, "id": identifier, "profit": round(profit, 2)})
    return normalized


def match_id_profit_rule(identifier_values: dict, rules: list[dict]) -> Optional[dict]:
    indexed = {
        (rule["id_type"], _identifier_key(rule["id"])): rule
        for rule in rules
    }
    for id_type in ID_RULE_TYPES:
        identifier = _identifier_key(identifier_values.get(id_type))
        if not identifier:
            continue
        rule = indexed.get((id_type, identifier))
        if rule:
            return {**rule, "matched_id": _text(identifier_values.get(id_type))}
    return None


def normalize_parse_config(config: object) -> Optional[dict]:
    if config is None:
        return None
    if not isinstance(config, dict):
        raise ValueError("SKC识别规则格式不正确")

    raw_keywords = config.get("set_keywords", [])
    if not isinstance(raw_keywords, list) or len(raw_keywords) > 20:
        raise ValueError("套装标识最多可设置20个")
    keywords = []
    for item in raw_keywords:
        keyword = _text(item)
        if len(keyword) > 32:
            raise ValueError("单个套装标识不能超过32个字符")
        if keyword not in keywords:
            keywords.append(keyword)

    raw_mappings = config.get("set_mappings", [])
    if not isinstance(raw_mappings, list) or len(raw_mappings) > 50:
        raise ValueError("套装固定映射最多可设置50个")
    mappings = []
    for item in raw_mappings:
        if not isinstance(item, dict):
            raise ValueError("套装固定映射格式不正确")
        pattern = _text(item.get("pattern"))
        if not pattern or len(pattern) > 64:
            raise ValueError("套装固定映射内容不能为空且不能超过64个字符")
        try:
            pieces = int(item.get("pieces"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"固定映射“{pattern}”的件数不正确") from exc
        if pieces not in ACTIVITY_PRICE_BASE:
            raise ValueError(f"固定映射“{pattern}”的{pieces}件套尚未配置活动价")
        mappings.append({"pattern": pattern, "pieces": pieces})

    single_mode = _text(config.get("single_mode")) or "last_segment"
    if single_mode not in SINGLE_PARSE_MODES:
        raise ValueError("单品货值提取方式不正确")
    delimiter = _text(config.get("single_delimiter"))
    marker = _text(config.get("single_marker"))
    if single_mode in {"first_segment", "last_segment"} and (not delimiter or len(delimiter) > 10):
        raise ValueError("单品分隔符不能为空且不能超过10个字符")
    if single_mode == "after_marker" and (not marker or len(marker) > 32):
        raise ValueError("单品指定文字不能为空且不能超过32个字符")

    return {
        "set_keywords": keywords,
        "set_mappings": mappings,
        "single_mode": single_mode,
        "single_delimiter": delimiter,
        "single_marker": marker,
    }


def parse_skc_detail(skc: object, parse_config: Optional[dict] = None, *, normalized: bool = False) -> Optional[dict]:
    value = _text(skc)
    if not value:
        return None

    if parse_config is not None:
        config = parse_config if normalized else normalize_parse_config(parse_config)
        folded_value = value.casefold()

        for mapping in config["set_mappings"]:
            if mapping["pattern"].casefold() in folded_value:
                return {"kind": "set", "value": float(mapping["pieces"]), "method": f"固定映射：{mapping['pattern']}"}

        for keyword in config["set_keywords"]:
            if keyword:
                match = re.search(rf"(\d+)\s*{re.escape(keyword)}", value, re.IGNORECASE)
                method = f"套装标识：{keyword}"
            else:
                match = re.search(r"(\d+)\s*$", value)
                method = "套装标识：空（末尾数字）"
            if match:
                pieces = int(match.group(1))
                if pieces not in ACTIVITY_PRICE_BASE:
                    return None
                return {"kind": "set", "value": float(pieces), "method": method}

        mode = config["single_mode"]
        candidate = ""
        method = ""
        if mode == "first_segment":
            delimiter = config["single_delimiter"]
            candidate = value.split(delimiter, 1)[0] if delimiter in value else ""
            method = f"第一个“{delimiter}”前的数字"
        elif mode == "last_segment":
            delimiter = config["single_delimiter"]
            candidate = value.rsplit(delimiter, 1)[-1] if delimiter in value else ""
            method = f"最后一个“{delimiter}”后的数字"
        else:
            marker = config["single_marker"]
            match = re.search(rf"{re.escape(marker)}\s*([0-9]+(?:\.[0-9]+)?)", value, re.IGNORECASE)
            candidate = match.group(1) if match else ""
            method = f"文字“{marker}”后的数字"

        if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", candidate.strip()):
            return {"kind": "single", "value": float(candidate), "method": method}
        return None

    set_match = SET_SKC_RE.fullmatch(value)
    if set_match:
        pieces = int(set_match.group(1))
        if pieces not in ACTIVITY_PRICE_BASE:
            return None
        return {"kind": "set", "value": float(pieces), "method": "系统套装格式"}

    single_match = SINGLE_SKC_RE.search(value)
    if single_match:
        return {"kind": "single", "value": float(single_match.group(1)), "method": "系统单品格式"}
    return None


def parse_skc(skc: object, parse_config: Optional[dict] = None, *, normalized: bool = False) -> Optional[Tuple[str, float]]:
    detail = parse_skc_detail(skc, parse_config, normalized=normalized)
    if not detail:
        return None
    return detail["kind"], detail["value"]


def activity_base_price(parsed: Tuple[str, float], settings=None, profit_adjustment: float = 0) -> float:
    kind, value = parsed
    activity_settings = (settings or {}).get("activity", {})
    if kind == "set":
        set_prices = activity_settings.get("set_prices", {})
        base = float(set_prices.get(str(int(value)), ACTIVITY_PRICE_BASE[int(value)]))
        return base + float(profit_adjustment)
    tiers = activity_settings.get("single_tiers", [{"min_price": 0, "profit": 0}])
    profit = max((float(tier.get("profit", 0)) for tier in tiers if value >= float(tier.get("min_price", 0))), default=0)
    return value + float(activity_settings.get("headcost", 5)) + float(activity_settings.get("operation_fee", 7)) + profit + float(profit_adjustment)


def _reference_price(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    try:
        price = float(value)
    except (TypeError, ValueError):
        return None
    return price if price >= 0 else None


def _uplifted_price(base: float, reference: float, uplift_limit: float) -> float:
    max_cents = min(int(round(max(0, uplift_limit) * 100)), int(round((reference - base) * 100)))
    if max_cents <= 0:
        return round(base, 2)
    uplift_cents = random.randint(1, max_cents)
    return round(base + uplift_cents / 100, 2)


def preview_activity_workbook(source: bytes, parse_config: Optional[dict] = None, settings=None, id_profit_rules: Optional[list[dict]] = None) -> dict:
    if not source:
        raise ValueError("上传的报名表为空")
    workbook = load_workbook(io.BytesIO(source), data_only=False, read_only=True, keep_links=False)
    try:
        worksheet, header_row, columns = _find_headers(workbook)
        activity_settings = (settings or {}).get("activity", {})
        effective_parse_config = parse_config if parse_config is not None else activity_settings.get("default_skc_rules")
        normalized = normalize_parse_config(effective_parse_config)
        effective_id_rules = normalize_id_profit_rules(
            id_profit_rules if id_profit_rules is not None else activity_settings.get("id_profit_rules", [])
        )
        skc_column = columns["skc"]
        items = []
        total = single_rows = set_rows = unrecognized_rows = id_rule_matches = 0
        for row in range(header_row + 1, worksheet.max_row + 1):
            skc = worksheet.cell(row, skc_column).value
            if not _text(skc):
                continue
            total += 1
            detail = parse_skc_detail(skc, normalized, normalized=True)
            if detail:
                kind = detail["kind"]
                if kind == "set":
                    set_rows += 1
                else:
                    single_rows += 1
                result = "套装" if kind == "set" else "单品"
                identifiers = {
                    id_type: worksheet.cell(row, columns[id_key]).value
                    for id_type, id_key in (("SPU", "spu_id"), ("SKC", "skc_id"), ("SKU", "sku_id"))
                    if id_key in columns
                }
                matched_rule = match_id_profit_rule(identifiers, effective_id_rules)
                adjustment = float(matched_rule["profit"]) if matched_rule else 0
                if matched_rule:
                    id_rule_matches += 1
                base_price = activity_base_price((kind, detail["value"]), settings)
                adjusted_price = base_price + adjustment
                item = {
                    "row": row,
                    "skc": _text(skc),
                    "spu_id": _text(identifiers.get("SPU")) or None,
                    "skc_id": _text(identifiers.get("SKC")) or None,
                    "sku_id": _text(identifiers.get("SKU")) or None,
                    "result": result,
                    "value": detail["value"],
                    "base_price": round(base_price, 2),
                    "adjusted_price": round(adjusted_price, 2),
                    "profit_adjustment": round(adjustment, 2),
                    "matched_id_type": matched_rule["id_type"] if matched_rule else None,
                    "matched_id": matched_rule["matched_id"] if matched_rule else None,
                    "method": detail["method"],
                }
            else:
                unrecognized_rows += 1
                item = {"row": row, "skc": _text(skc), "result": "无法识别", "value": None, "base_price": None, "method": "未匹配规则或套装件数未配置"}
            if len(items) < MAX_PREVIEW_ROWS:
                items.append(item)
        return {
            "sheet": worksheet.title,
            "header_row": header_row,
            "total_rows": total,
            "single_rows": single_rows,
            "set_rows": set_rows,
            "unrecognized_rows": unrecognized_rows,
            "id_profit_rule_matches": id_rule_matches,
            "preview_limit": MAX_PREVIEW_ROWS,
            "items": items,
        }
    finally:
        workbook.close()


def process_activity_workbook(source: bytes, output_path: Path, settings=None, parse_config: Optional[dict] = None, id_profit_rules: Optional[list[dict]] = None) -> dict:
    if not source:
        raise ValueError("上传的报名表为空")

    workbook = load_workbook(io.BytesIO(source), data_only=False, keep_links=False)
    try:
        worksheet, header_row, columns = _find_headers(workbook)
        activity_settings = (settings or {}).get("activity", {})
        custom_parse_config = parse_config is not None
        effective_parse_config = parse_config if custom_parse_config else activity_settings.get("default_skc_rules")
        normalized_parse_config = normalize_parse_config(effective_parse_config)
        effective_id_rules = normalize_id_profit_rules(
            id_profit_rules if id_profit_rules is not None else activity_settings.get("id_profit_rules", [])
        )
        price_column = columns["price"]
        skc_column = columns["skc"]
        uplift_limit = float(activity_settings.get("uplift_limit", 1))
        rows_to_delete = []
        updated = 0
        unchanged = 0
        skipped = 0
        processed = 0
        id_rule_matches = 0
        input_data_rows = max(0, worksheet.max_row - header_row)

        for row in range(header_row + 1, worksheet.max_row + 1):
            skc = worksheet.cell(row, skc_column).value
            if not _text(skc):
                skipped += 1
                continue
            parsed = parse_skc(skc, normalized_parse_config, normalized=True)
            reference = _reference_price(worksheet.cell(row, price_column).value)
            if parsed is None or reference is None:
                skipped += 1
                continue

            processed += 1
            identifiers = {
                id_type: worksheet.cell(row, columns[id_key]).value
                for id_type, id_key in (("SPU", "spu_id"), ("SKC", "skc_id"), ("SKU", "sku_id"))
                if id_key in columns
            }
            matched_rule = match_id_profit_rule(identifiers, effective_id_rules)
            adjustment = float(matched_rule["profit"]) if matched_rule else 0
            if matched_rule:
                id_rule_matches += 1
            base = activity_base_price(parsed, settings, adjustment)
            if reference < base:
                rows_to_delete.append(row)
                continue
            if abs(reference - base) < 0.000001:
                unchanged += 1
                continue

            worksheet.cell(row, price_column).value = _uplifted_price(base, reference, uplift_limit)
            worksheet.cell(row, price_column).number_format = "0.00"
            updated += 1

        for row in reversed(rows_to_delete):
            worksheet.delete_rows(row, 1)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return {
            "sheet": worksheet.title,
            "header_row": header_row,
            "input_data_rows": input_data_rows,
            "processed_rows": processed,
            "updated_rows": updated,
            "unchanged_rows": unchanged,
            "removed_rows": len(rows_to_delete),
            "skipped_rows": skipped,
            "remaining_data_rows": input_data_rows - len(rows_to_delete),
            "uplift_limit": round(uplift_limit, 2),
            "custom_skc_rules": custom_parse_config,
            "id_profit_rule_matches": id_rule_matches,
            "custom_id_profit_rules": id_profit_rules is not None,
        }
    finally:
        workbook.close()
