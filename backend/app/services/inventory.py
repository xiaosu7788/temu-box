from __future__ import annotations

import io
import json
import os
import re
import threading
import time
from pathlib import Path
from typing import Callable, Dict, Optional, Tuple, Union

from openpyxl import load_workbook

from app.config import INVENTORY_PATH, PRICE_CACHE_PATH
from app.database import (
    current_inventory_metadata,
    get_inventory_catalog,
    invalidate_inventory_catalog,
    inventory_signature_matches,
    save_inventory_catalog,
)


LogFn = Optional[Callable[[str], None]]
WorkbookInput = Union[str, Path, bytes, bytearray]

CODE_RE = re.compile(r"MB131-[A-Za-z0-9]+")
PRICE_HEADER_RE = re.compile(r"(价格|货值|单价|售价|成本价|price)", re.IGNORECASE)
SET_RE = re.compile(
    r"(十二|十一|十|九|八|七|六|五|四|三|二|一|12|11|10|9|8|7|6|5|4|3|2|1)\s*件套"
)
SET_NUM_MAP = {
    "一": 1, "1": 1, "二": 2, "2": 2, "三": 3, "3": 3,
    "四": 4, "4": 4, "五": 5, "5": 5, "六": 6, "6": 6,
    "七": 7, "7": 7, "八": 8, "8": 8, "九": 9, "9": 9,
    "十": 10, "10": 10, "十一": 11, "11": 11, "十二": 12, "12": 12,
}

SHEET_RULES = {
    "Sheet1": {3: 6, 10: 8},
    "美东": {3: 5, 5: 3, 8: 9, 9: 8},
    "加拿大1仓": {10: 8, 6: 3},
    "加拿大2仓（美国S仓）": {10: 8, 3: 6},
    "美西2仓": {8: 9, 4: 3},
    "澳-日-英": {12: 10},
    "内衣": {5: 3},
    "cos，丝袜": {4: 5},
    "玩具": {4: 6},
    "个人单品": {3: 6},
    "库存0": {3: 6, 4: 5, 5: 3, 6: 8, 7: 6, 8: 9, 9: 7, 10: 8},
}
PRIORITY = [
    "Sheet1", "美东", "加拿大1仓", "加拿大2仓（美国S仓）", "美西2仓",
    "澳-日-英", "内衣", "cos，丝袜", "玩具", "个人单品", "库存0",
]

PRICE_CACHE_VERSION = 3
_CACHE_LOCK = threading.Lock()


def _log(log: LogFn, message: str) -> None:
    if log:
        log(message)


def to_float(value: object) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def to_price(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("=DISPIMG") or CODE_RE.search(text):
            return None
        text = text.replace(",", "")
        match = re.fullmatch(r"[￥¥$€]?\s*(\d+(?:\.\d+)?)\s*(?:元|块)?", text)
        if not match:
            return None
        number = float(match.group(1))
    else:
        number = to_float(value)
    if number is None or not 0 < number < 1000:
        return None
    return number


def extract_set_type(row_texts) -> Optional[str]:
    for text in row_texts:
        if not text:
            continue
        for match in SET_RE.finditer(str(text)):
            number = SET_NUM_MAP.get(match.group(1))
            if number:
                return f"{number}件套"
        if re.search(r"单品", str(text)):
            return "单品"
        if re.search(r"多件套|套装|组合", str(text)):
            return "多件套"
    return None


def build_price_header_index(ws, max_col: int = 20):
    """Record the nearest price header group above every worksheet row."""
    limit = min(ws.max_column, max_col)
    index = [set() for _ in range(ws.max_row + 1)]
    active_columns = set()
    for row in range(1, ws.max_row + 1):
        index[row] = set(active_columns)
        row_columns = set()
        for column in range(1, limit + 1):
            value = ws.cell(row, column).value
            if value is not None and PRICE_HEADER_RE.search(str(value).strip()):
                row_columns.add(column)
        if row_columns:
            active_columns = row_columns
    return index


def find_row_price(
    ws,
    row: int,
    mapped_price_col: int,
    sku_col: Optional[int] = None,
    price_header_cols=None,
    max_col: int = 20,
) -> Tuple[Optional[float], Optional[int]]:
    """Find price by nearest header, legacy mapping, then expanding around SKU."""
    limit = min(ws.max_column, max_col)
    if price_header_cols:
        header_col = min(
            (col for col in price_header_cols if col <= limit),
            key=lambda col: (abs(col - sku_col), col) if sku_col else (col, col),
            default=None,
        )
        if header_col is not None:
            price = to_price(ws.cell(row, header_col).value)
            if price is not None:
                return price, header_col

    price = to_price(ws.cell(row, mapped_price_col).value)
    if price is not None:
        return price, mapped_price_col

    if sku_col is not None:
        for distance in range(1, limit + 1):
            for column in (sku_col - distance, sku_col + distance):
                if column < 1 or column > limit or column == mapped_price_col:
                    continue
                price = to_price(ws.cell(row, column).value)
                if price is not None:
                    return price, column
    return None, None


def build_price_catalog(source: WorkbookInput, log: LogFn = None) -> Dict[str, dict]:
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True, keep_links=False)
    else:
        wb = load_workbook(source, data_only=True, keep_links=False)

    candidates: Dict[str, list] = {}
    try:
        for priority, sheet_name in enumerate(PRIORITY):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rules = SHEET_RULES.get(sheet_name, {})
            if not rules:
                continue
            header_index = build_price_header_index(ws)
            for row in range(1, ws.max_row + 1):
                if not any(
                    isinstance(ws.cell(row, col).value, str)
                    and CODE_RE.search(ws.cell(row, col).value)
                    for col in rules
                ):
                    continue

                row_texts = None
                set_type = None
                for sku_col, mapped_price_col in rules.items():
                    value = ws.cell(row, sku_col).value
                    if not isinstance(value, str):
                        continue
                    cell_text = value.strip()
                    codes = CODE_RE.findall(cell_text)
                    if not codes:
                        continue
                    if row_texts is None:
                        row_texts = [
                            str(ws.cell(row, col).value)
                            for col in range(1, min(ws.max_column, 20) + 1)
                            if ws.cell(row, col).value is not None
                            and not str(ws.cell(row, col).value).startswith("=DISPIMG")
                        ]
                        set_type = extract_set_type(row_texts) or "单品"
                    price, actual_price_col = find_row_price(
                        ws, row, mapped_price_col, sku_col, header_index[row]
                    )
                    if price is None:
                        continue
                    for code in codes:
                        candidates.setdefault(code, []).append({
                            "sku": code,
                            "price": price,
                            "set_type": set_type,
                            "source_sheet": sheet_name,
                            "source_row": row,
                            "source_column": actual_price_col,
                            "priority": priority,
                        })
    finally:
        wb.close()

    catalog = {}
    for sku, sku_candidates in candidates.items():
        catalog[sku] = min(
            sku_candidates,
            key=lambda item: (item["price"], item["priority"]),
        )
        catalog[sku].pop("priority", None)
    _log(log, f"库存扫描完成，共提取 {len(catalog)} 个 SKU，重复 SKU 取最低货值。")
    return catalog


def inventory_signature(path: Path = INVENTORY_PATH) -> dict:
    stat = path.stat()
    return {
        "path": str(path.resolve()),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "parser_version": PRICE_CACHE_VERSION,
    }


def invalidate_cache() -> None:
    invalidate_inventory_catalog()
    try:
        PRICE_CACHE_PATH.unlink()
    except FileNotFoundError:
        pass


def load_price_catalog(path: Path = INVENTORY_PATH, log: LogFn = None) -> Dict[str, dict]:
    if not path.exists():
        catalog = get_inventory_catalog() if path == INVENTORY_PATH else {}
        if catalog:
            _log(log, f"库存原始表暂不可用，使用数据库库存数据，共 {len(catalog)} 个 SKU。")
            return catalog
        raise FileNotFoundError(f"库存统计表不存在：{path}")
    signature = inventory_signature(path)
    with _CACHE_LOCK:
        if path == INVENTORY_PATH and inventory_signature_matches(signature):
            catalog = get_inventory_catalog()
            if catalog:
                _log(log, f"使用数据库库存数据，共 {len(catalog)} 个 SKU。")
                return catalog
        try:
            with PRICE_CACHE_PATH.open("r", encoding="utf-8") as handle:
                cached = json.load(handle)
            if cached.get("signature") == signature:
                catalog = cached.get("catalog", {})
                if path == INVENTORY_PATH and catalog:
                    save_inventory_catalog(signature, catalog)
                _log(log, f"使用库存缓存，共 {len(catalog)} 个 SKU。")
                return catalog
        except (FileNotFoundError, json.JSONDecodeError, OSError, TypeError):
            pass

        started = time.time()
        _log(log, "库存缓存无效，正在重新扫描库存统计表...")
        catalog = build_price_catalog(path, log)
        if path == INVENTORY_PATH:
            save_inventory_catalog(signature, catalog)
        payload = {
            "signature": signature,
            "catalog": catalog,
            "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        PRICE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        temp_path = PRICE_CACHE_PATH.with_suffix(PRICE_CACHE_PATH.suffix + ".tmp")
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False)
        os.replace(temp_path, PRICE_CACHE_PATH)
        _log(log, f"库存缓存已更新，耗时 {time.time() - started:.1f} 秒。")
        return catalog


def inventory_status() -> dict:
    metadata = current_inventory_metadata()
    status = {
        "path": str(INVENTORY_PATH),
        "exists": INVENTORY_PATH.exists(),
        "cache_exists": PRICE_CACHE_PATH.exists(),
        "legacy_cache_valid": False,
        "sku_count": metadata["sku_count"] if metadata else 0,
        "size": None,
        "modified_at": None,
        "cache_valid": bool(metadata and INVENTORY_PATH.exists() and inventory_signature(INVENTORY_PATH) == {key: metadata[key] for key in ("path", "size", "mtime_ns", "parser_version")}),
        "database": {
            "configured": True,
            "has_inventory": metadata is not None,
        },
    }
    if INVENTORY_PATH.exists():
        stat = INVENTORY_PATH.stat()
        status["size"] = stat.st_size
        status["modified_at"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime))
    try:
        with PRICE_CACHE_PATH.open("r", encoding="utf-8") as handle:
            cached = json.load(handle)
        status["legacy_cache_valid"] = (
            INVENTORY_PATH.exists()
            and cached.get("signature") == inventory_signature(INVENTORY_PATH)
        )
    except (FileNotFoundError, json.JSONDecodeError, OSError, TypeError):
        pass
    return status
