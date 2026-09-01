from __future__ import annotations

import io
from pathlib import Path
from typing import Callable, Dict, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.inventory import to_float


WorkbookInput = Union[str, Path, bytes, bytearray]
LogFn = Optional[Callable[[str], None]]

HEADCOST_MAP = {
    "单品": 5,
    "4件套": 5,
    "5件套": 5,
    "6件套": 5,
    "8件套": 10,
    "10件套": 10,
    "12件套": 15,
}
OPERATION_FEE = 7
EXTRA_ITEM_FEE = 2


def calc_order_cost(sku_items, half_headcost_skus=None):
    if not sku_items:
        return None
    half_headcost_skus = half_headcost_skus or {}
    total = 0.0
    total_qty = 0
    for price, set_type, quantity, sku in sku_items:
        if price is None:
            return None
        headcost = HEADCOST_MAP.get(set_type, 5)
        if sku in half_headcost_skus:
            headcost /= 2
        total += quantity * (price + headcost)
        total_qty += quantity
    total += OPERATION_FEE
    if total_qty > 1:
        total += (total_qty - 1) * EXTRA_ITEM_FEE
    return round(total, 2)


def _find_header_row(ws, keywords, max_scan=10):
    for row in range(1, min(max_scan + 1, ws.max_row + 1)):
        found = {}
        for column in range(1, min(ws.max_column, 40) + 1):
            value = ws.cell(row, column).value
            if value is None:
                continue
            text = str(value).strip()
            for keyword in keywords:
                if keyword in text and keyword not in found:
                    found[keyword] = column
        if all(keyword in found for keyword in keywords):
            return row, found
    return None, {}


def build_delivery_sku_map(source: WorkbookInput, log: LogFn = None):
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True, keep_links=False)
    else:
        wb = load_workbook(source, data_only=True, keep_links=False)
    try:
        ws = wb.active
        header_row, columns = _find_header_row(ws, ["参考单号", "SKU", "数量"])
        if header_row is None:
            ref_col, sku_col, qty_col, data_start = 2, 20, 21, 2
            if log:
                log("派送订单未识别到表头，已使用兼容列配置。")
        else:
            ref_col = columns["参考单号"]
            sku_col = columns["SKU"]
            qty_col = columns["数量"]
            data_start = header_row + 1

        result = {}
        current_ref = None
        for row in range(data_start, ws.max_row + 1):
            ref = ws.cell(row, ref_col).value
            sku = ws.cell(row, sku_col).value
            quantity = ws.cell(row, qty_col).value
            if ref and str(ref).strip():
                current_ref = str(ref).strip()
            if not sku or not str(sku).strip() or not current_ref:
                continue
            sku_text = str(sku).strip()
            qty_value = to_float(quantity)
            if qty_value is None or qty_value < 1:
                qty_value = 1
            entries = result.setdefault(current_ref, [])
            for index, (existing_sku, existing_qty) in enumerate(entries):
                if existing_sku == sku_text:
                    entries[index] = (existing_sku, existing_qty + int(qty_value))
                    break
            else:
                entries.append((sku_text, int(qty_value)))
        if log:
            log(f"派送订单扫描完成，共 {len(result)} 个 PO 单号。")
        return result
    finally:
        wb.close()


def generate_summary(
    sales_source: WorkbookInput,
    catalog: Dict[str, dict],
    po_to_skus,
    output_path: Path,
    half_headcost_skus=None,
    log: LogFn = None,
):
    half_headcost_skus = half_headcost_skus or {}
    if isinstance(sales_source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(sales_source), data_only=True, keep_links=False)
    else:
        wb = load_workbook(sales_source, data_only=True, keep_links=False)
    try:
        ws = wb.active
        header_row, columns = _find_header_row(ws, ["PO单号", "结算金额"])
        if header_row is None:
            po_col, amount_col, data_start = 3, 6, 3
        else:
            po_col = columns["PO单号"]
            amount_col = columns["结算金额"]
            data_start = header_row + 1

        results = []
        matched = 0
        unmatched = set()
        missing_skus = set()
        type_stats = {}
        for row in range(data_start, ws.max_row + 1):
            po_value = ws.cell(row, po_col).value
            if not po_value or not str(po_value).strip():
                continue
            po = str(po_value).strip()
            amount = to_float(ws.cell(row, amount_col).value)
            sku_qty_list = po_to_skus.get(po, [])
            if not sku_qty_list:
                unmatched.add(po)
                results.append((po, "", None, 0, None, amount, ""))
                continue

            matched += 1
            prices = []
            types = []
            skus = []
            total_qty = 0
            cost_items = []
            for sku, quantity in sku_qty_list:
                item = catalog.get(sku, {})
                price = item.get("price")
                set_type = half_headcost_skus.get(sku) or item.get("set_type", "单品")
                if price is None:
                    missing_skus.add(sku)
                prices.append(price)
                types.append(set_type)
                skus.append(sku)
                total_qty += quantity
                cost_items.append((price, set_type, quantity, sku))

            price_display = prices[0] if len(prices) == 1 else "/".join(
                f"{price:.2f}" if price is not None else "" for price in prices
            )
            unique_types = list(dict.fromkeys(types))
            for set_type in unique_types:
                type_stats[set_type] = type_stats.get(set_type, 0) + 1
            results.append((
                po,
                "/".join(skus),
                price_display,
                total_qty,
                calc_order_cost(cost_items, half_headcost_skus),
                amount,
                "/".join(unique_types),
            ))
    finally:
        wb.close()

    output = Workbook()
    out_ws = output.active
    out_ws.title = "汇总表"
    headers = ["PO单号", "SKU", "货值", "件数", "成本", "结算金额", "类型"]
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(*[Side(style="thin", color="D1D5DB")] * 4)
    for column, header in enumerate(headers, 1):
        cell = out_ws.cell(1, column, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_number, row_data in enumerate(results, 2):
        for column, value in enumerate(row_data, 1):
            cell = out_ws.cell(row_number, column, value)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = center
            cell.border = border
            if column in (3, 5, 6) and value is not None and not isinstance(value, str):
                cell.number_format = "0.00"

    merge_columns = [1, 2, 3, 4, 5, 7]
    start = 2
    while start <= len(results) + 1:
        end = start
        po = results[start - 2][0]
        while end <= len(results) + 1 and results[end - 2][0] == po:
            end += 1
        if end - start > 1:
            for column in merge_columns:
                out_ws.merge_cells(start_row=start, start_column=column, end_row=end - 1, end_column=column)
        start = end

    for column, width in enumerate([29, 32, 11, 9, 12, 14, 12], 1):
        out_ws.column_dimensions[chr(64 + column)].width = width
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = out_ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    output.close()

    stats = {
        "total": len(results),
        "matched": matched,
        "unmatched": len(unmatched),
        "unmatched_pos": sorted(unmatched),
        "missing_skus": sorted(missing_skus),
        "type_stats": type_stats,
    }
    if log:
        log(
            f"汇总表生成完成：{len(results)} 行，匹配 {matched} 行，"
            f"未匹配 {len(unmatched)} 行。"
        )
    return stats
